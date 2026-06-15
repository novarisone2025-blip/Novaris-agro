import io
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.graphics import renderSVG
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.analytics import (
    animal_dict,
    arroba_summary,
    finance_summary,
    get_animals,
    health_dict,
    lot_summary,
    paddock_dict,
    reproduction_dict,
    weighing_dict,
)
from app.models import (
    Animal,
    FinancialEntry,
    HealthRecord,
    Paddock,
    ReproductionEvent,
    User,
    Weighing,
)


REPORT_NAMES = {
    "rebanho": "Relatório do rebanho",
    "sanitario": "Relatório sanitário",
    "financeiro": "Relatório financeiro",
    "reprodutivo": "Relatório reprodutivo",
    "pesagens": "Relatório de pesagens",
    "arrobas": "Relatório de arrobas",
    "lotes": "Relatório de lotes",
    "pastagens": "Relatório de pastagens",
}


def report_data(kind: str, db: Session, user: User) -> tuple[list[str], list[list]]:
    animals = get_animals(db, user.farm_id)
    animal_map = {item.id: item for item in animals}
    if kind == "rebanho":
        return (
            ["Brinco", "Nome", "Raça", "Sexo", "Peso", "Lote", "Piquete", "Status"],
            [
                [
                    item.tag_number,
                    item.name or "",
                    item.breed,
                    item.sex,
                    item.current_weight,
                    item.lot,
                    item.paddock,
                    item.status,
                ]
                for item in animals
            ],
        )
    if kind == "sanitario":
        records = db.scalars(
            select(HealthRecord)
            .join(Animal, Animal.id == HealthRecord.animal_id)
            .where(Animal.farm_id == user.farm_id)
            .order_by(HealthRecord.applied_at.desc())
        ).all()
        return (
            ["Animal", "Tipo", "Produto", "Aplicação", "Próxima", "Lote"],
            [
                [
                    animal_map[item.animal_id].tag_number,
                    item.record_type,
                    item.product_name,
                    item.applied_at,
                    item.next_application_at or "",
                    item.batch or "",
                ]
                for item in records
            ],
        )
    if kind == "financeiro":
        entries = db.scalars(
            select(FinancialEntry)
            .where(FinancialEntry.farm_id == user.farm_id)
            .order_by(FinancialEntry.occurred_at.desc())
        ).all()
        return (
            ["Data", "Tipo", "Categoria", "Descrição", "Valor", "Lote"],
            [
                [
                    item.occurred_at,
                    item.entry_type,
                    item.category,
                    item.description,
                    item.amount,
                    item.lot or "",
                ]
                for item in entries
            ],
        )
    if kind == "reprodutivo":
        events = db.scalars(
            select(ReproductionEvent)
            .join(Animal, Animal.id == ReproductionEvent.animal_id)
            .where(Animal.farm_id == user.farm_id)
            .order_by(ReproductionEvent.event_date.desc())
        ).all()
        return (
            ["Animal", "Evento", "Data", "Resultado", "Parto previsto"],
            [
                [
                    animal_map[item.animal_id].tag_number,
                    item.event_type,
                    item.event_date,
                    item.result or "",
                    item.expected_calving_at or "",
                ]
                for item in events
            ],
        )
    if kind == "pesagens":
        items = db.scalars(
            select(Weighing)
            .join(Animal, Animal.id == Weighing.animal_id)
            .where(Animal.farm_id == user.farm_id)
            .order_by(Weighing.weighed_at.desc())
        ).all()
        return (
            ["Animal", "Data", "Peso (kg)", "Observações"],
            [
                [
                    animal_map[item.animal_id].tag_number,
                    item.weighed_at,
                    item.weight,
                    item.notes or "",
                ]
                for item in items
            ],
        )
    if kind == "arrobas":
        rows = arroba_summary(db, user)["animals"]
        return (
            ["Animal", "Lote", "Peso vivo", "Carcaça", "Arrobas", "Valor estimado"],
            [
                [
                    item["tag_number"],
                    item["lot"],
                    item["live_weight"],
                    item["carcass_weight"],
                    item["estimated_arrobas"],
                    item["estimated_value"],
                ]
                for item in rows
            ],
        )
    if kind == "lotes":
        rows = lot_summary(db, user)["items"]
        return (
            ["Lote", "Animais", "Peso médio", "GMD", "Receita", "Despesas", "Lucro"],
            [
                [
                    item["lot"],
                    item["animal_count"],
                    item["average_weight"],
                    item["average_daily_gain"] or "",
                    item["revenue"],
                    item["expenses"],
                    item["estimated_profit"],
                ]
                for item in rows
            ],
        )
    if kind == "pastagens":
        rows = db.scalars(
            select(Paddock)
            .where(Paddock.farm_id == user.farm_id)
            .order_by(Paddock.name)
        ).all()
        return (
            ["Piquete", "Área (ha)", "Capacidade", "Animais", "Ocupação", "Status"],
            [
                [
                    item.name,
                    item.area_hectares,
                    item.capacity,
                    item.current_animals,
                    paddock_dict(item)["occupancy_rate"],
                    item.status,
                ]
                for item in rows
            ],
        )
    raise ValueError("Relatório desconhecido.")


def generate_pdf(kind: str, db: Session, user: User) -> bytes:
    headers, rows = report_data(kind, db, user)
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("NOVARIS AGRO", styles["Title"]),
        Paragraph(escape(REPORT_NAMES[kind]), styles["Heading2"]),
        Paragraph(
            escape(f"{user.farm.name} • {user.farm.city}/{user.farm.state}"),
            styles["Normal"],
        ),
        Spacer(1, 7 * mm),
    ]
    table_data = [headers] + [
        ["" if value is None else str(value) for value in row] for row in rows
    ]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#245f46")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d8e2dc")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f8f6")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return output.getvalue()


def generate_xlsx(kind: str, db: Session, user: User) -> bytes:
    headers, rows = report_data(kind, db, user)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"
    sheet.append([REPORT_NAMES[kind]])
    sheet.append([f"{user.farm.name} - {user.farm.city}/{user.farm.state}"])
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[4]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="245F46")
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2,
            42,
        )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_qr_svg(content: str) -> bytes:
    widget = QrCodeWidget(content)
    bounds = widget.getBounds()
    size = 120
    drawing = Drawing(size, size)
    drawing.add(widget)
    widget.width = size
    widget.height = size
    widget.x = -bounds[0]
    widget.y = -bounds[1]
    return renderSVG.drawToString(drawing).encode("utf-8")


def generate_animal_pdf(animal: Animal, db: Session) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    data = animal_dict(animal)
    rows = [
        ["Brinco", data["tag_number"]],
        ["Nome", data["name"] or "Não informado"],
        ["Raça / sexo", f"{data['breed']} / {data['sex']}"],
        ["Nascimento", str(data["birth_date"])],
        ["Peso atual", f"{data['current_weight']} kg"],
        ["Lote / piquete", f"{data['lot']} / {data['paddock']}"],
        ["Status", data["status"]],
        ["Código Novaris", data["unique_code"] or ""],
        ["Pai / mãe", f"{data['father_tag'] or '-'} / {data['mother_tag'] or '-'}"],
    ]
    table = Table(rows, colWidths=[45 * mm, 110 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8f1ec")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd8d0")),
                ("PADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    document.build(
        [
            Paragraph("NOVARIS AGRO", styles["Title"]),
            Paragraph("Ficha individual do animal", styles["Heading2"]),
            Spacer(1, 7 * mm),
            table,
        ]
    )
    return output.getvalue()
